"""Parser de rodas (Cheney, ES2WHEELS — mesmo formato XLSX). Formato
conhecido (arquivos reais inspecionados, 2026-08-12): sem preço, sem
compatibilidade de veículo (compatibilidade é por medida — PCD/ET/CB, não
marca/modelo de carro). BRAND só aparece na primeira linha de cada grupo
(preenchido pra baixo nas linhas seguintes até a próxima marca).

Sem coluna de referência/SKU única — construída a partir de
modelo+medida+furação+offset+acabamento (única combinação que identifica
uma variante real do produto nesses arquivos).

Preço: nenhum nos arquivos recebidos — `cost_price` fica None (produto
aparece como "preço indisponível" até o cliente responder
perguntas-para-cliente.txt, 7.3).
"""

import io

import openpyxl

from .base import ParsedProduct, sanitize_text

HEADER_ROW_MARKERS = {"BRAND", "PHOTO"}


def parse_wheels_xlsx(file_bytes: bytes) -> dict[str, ParsedProduct]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    products: dict[str, ParsedProduct] = {}
    current_brand = ""
    header_found = False

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row and str(row[0] or "").strip().upper() == "BRAND":
                header_found = True
            continue

        if not row or all(v is None for v in row):
            continue

        brand, photo, model, size, pcd, et, cb, finish, stock = (row + (None,) * 9)[:9]

        if brand:
            current_brand = str(brand).strip()
        if not model:
            continue

        model = str(model).strip()
        size = str(size or "").strip()
        pcd = str(pcd or "").strip()
        finish_str = str(finish or "").strip()

        reference = "-".join(
            sanitize_text(v) for v in [model, size, pcd, str(et or ""), str(cb or ""), finish_str]
        )
        title = f"{current_brand} {model} {size}".strip()

        products[reference] = ParsedProduct(
            external_reference=reference,
            title=sanitize_text(title),
            description="",
            cost_price=None,
            image_url="",
            category=sanitize_text(current_brand),
            raw_attributes={
                "brand": current_brand,
                "model": model,
                "size": size,
                "pcd": pcd,
                "et": et,
                "cb": cb,
                "finish": finish_str,
                "stock": stock,
            },
        )

    wb.close()
    return products
